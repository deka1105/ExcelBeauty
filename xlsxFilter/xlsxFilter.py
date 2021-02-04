class xlsxFilter:
    import pandas as pd
    import numpy as np
    sourcefilePath = ""
    logfilePath = ""
    xlsxfilePath = ""
    indexfilePath = ""
    defcon = pd.DataFrame()

    def __init__(self, src, log, xls, ind):
        xlsxFilter.sourcefilePath = src
        xlsxFilter.logfilePath = log
        xlsxFilter.xlsxfilePath = xls
        xlsxFilter.indexfilePath = ind
        print("sourcefilePath: ", xlsxFilter.sourcefilePath)
        print("logfilePath   : ", xlsxFilter.logfilePath)
        print("xlsxfilePath  : ", xlsxFilter.xlsxfilePath)
        print("indexfilePath : ", xlsxFilter.indexfilePath)

    def initiate(self, worksheet):
        file = xlsxFilter.sourcefilePath
        #print(file)
        xl = xlsxFilter.pd.ExcelFile(file)
        #print(xl.sheet_names)
        xlsxFilter.defcon = xl.parse(worksheet)
        print(xlsxFilter.defcon.head())
        return xlsxFilter.defcon

    def reduceData(self, df1, column, paramsList, sortType):
        count = 0
        df = xlsxFilter.pd.DataFrame(columns=list(df1))
        print(column)
        print(paramsList)
        print("Input Count:", len(df1))
        for index in range (0, len(df1)):
            inputFlag = len(paramsList)
            listcount = 0
            for i in range(0, len(paramsList)):
                locationIndex = -1
                #str(df1[column].get(index))
                locationIndex = str(df1[column].get(index)).find(str(paramsList[i]))
                if(sortType == 'ex'):
                    if (locationIndex != -1):
                        inputFlag -= 1
                    if(i == len(paramsList)-1 and inputFlag == len(paramsList)):
                        inputFlag = 0
                if((sortType == 'in' and locationIndex != -1) or inputFlag == 0):
                    count += 1
                    ddf = df1.iloc[index].copy()
                    df = df.append(ddf, ignore_index=True)
                    break
        print("Output Count:", len(df))
        return df

    def nullRemover(self, df1, column):
        print("Input Rows:", len(df1))
        df = df1.dropna(subset = column)
        print("Output Rows:", len(df))
        return df

    def save_xls(self, list_dfs, xls_path):
        from pandas import ExcelWriter
        with ExcelWriter(xls_path) as writer:
            for n, df in enumerate(list_dfs):
                df.to_excel(writer,'sheet%s' % n)
            writer.save()

    def exportToExcel(self, filePath, dataFrameList):
        if(filePath == 'null'):
            import datetime
            import time
            ts = time.time()
            st = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d_%H-%M-%S')
            filePath = xlsxFilter.xlsxfilePath + st + ".xlsx"
        xlsxFilter.save_xls(self, list(dataFrameList), filePath)
        return filePath

    def line_prepender(self, filename, line):
        with open(filename, 'r+') as f:
            content = f.read()
            f.seek(0, 0)
            f.write(line.rstrip('\r\n') + '\n' + content)

    def addToLogIndex(self, qCode, queryText, fileLinktxt, fileLinkxlsx):
        import openpyxl
        filename = xlsxFilter.indexfilePath
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        data1 = "=HYPERLINK(\"" + fileLinktxt +"\", \"Logs txt File\")"
        data2 = "=HYPERLINK(\"" + fileLinkxlsx +"\", \"XLSX\")"
        rowNumber =  ws.max_row+1
        ws.cell(column=1, row=rowNumber, value=qCode)
        ws.cell(column=2, row=rowNumber, value=queryText)
        ws.cell(column=3, row=rowNumber, value=data1)
        ws.cell(column=4, row=rowNumber, value=data2)
        wb.save(filename)

    def FilterDataFrame(self, df1, column, paramsList):
        count = 0
        df = xlsxFilter.pd.DataFrame(columns=list(df1))
        for index in range (0, len(df1)):
            for i in range(0, len(paramsList)):
                locationIndex = -1
                temp_a = str(df1[column].get(index))
                temp_b = str(paramsList[i])
                locationIndex = temp_a.find(temp_b)
                if(locationIndex != -1):
                    count += 1
                    ddf = df1.iloc[index].copy()
                    df = df.append(ddf, ignore_index=True)
                    break
        return df

    def FilterDataFrameAdv(self, qCode, queryText, df1, primarycolumn, secondaryField,
                           exclusiveInclude, exclusiveExclude, exportFlag, xlsxFilePath):
        import datetime
        import time
        ts = time.time()
        st = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d_%H-%M-%S')
        timeStampFileName = st
        filePath = xlsxFilter.logfilePath + st + ".txt"
        f = open(filePath, "a")
        count = 0
        #df = list()
        df = xlsxFilter.pd.DataFrame(columns=list(df1))
        print("FileName:", st)
        print("FilePath:", filePath)
        #print("FdfAdv-In Function")
        print("FdfAdv--- on column --", primarycolumn)
        print("FdfAdv-Include only:", exclusiveInclude)
        print("FdfAdv-With Exclusion:", exclusiveExclude)
        print("FdfAdv-DataFrame Length:", len(df1))
        print("FdfAdv-", len(df))
        stripedData = df1[primarycolumn].unique()
        stripedData = list(stripedData)
        for stripIndex in range (0, len(stripedData)):
            flag = 0
            dfx = xlsxFilter.pd.DataFrame(columns=list(df1))
            dfy = xlsxFilter.pd.DataFrame(columns=list(df1))
            dfz = xlsxFilter.pd.DataFrame(columns=list(df1))
            dfxLength = 0
            dfyLength = 0
            dfzLength = 0
            arg = list()
            arg.append(stripedData[stripIndex])
            dfx = xlsxFilter.FilterDataFrame(self, df1, primarycolumn, arg)
            dfxLength = len(dfx)
            #print("dfx",primarycolumn, arg, dfxLength)
            f.write("dfx - "+str(primarycolumn)+"---"+str(arg)+"---"+str(dfxLength)+'\n')

            dfy = xlsxFilter.FilterDataFrame(self, dfx,  secondaryField, exclusiveInclude)
            dfyLength = len(dfy)
            #print("dfy", secondaryField, exclusiveInclude, dfyLength)
            f.write("dfy - "+str(secondaryField)+"---"+str(exclusiveInclude)+"---"+str(dfyLength)+'\n')

            dfz = xlsxFilter.FilterDataFrame(self, dfx,  secondaryField, exclusiveExclude)
            dfzLength = len(dfz)
            #print("dfz", secondaryField, exclusiveExclude, dfzLength)
            f.write("dfx - "+str(secondaryField)+"---"+str(exclusiveExclude)+"---"+str(dfzLength)+'\n')

            #print("dfx: ", dfxLength, "dfy: ", dfyLength, "dfz: ", dfzLength)
            f.write("dfx: "+str(dfxLength)+"dfy: "+str(dfyLength)+"dfz: "+str(dfzLength)+'\n')

            if(dfyLength > 0 and dfzLength == 0):
                #df.append(stripedData[stripIndex]+"-"+str(dfyLength))
                #df.loc[len(df)] = [stripedData[stripIndex],dfyLength]
                df = df.append(dfy)
        print("FdfAdv-", len(df))
        #final write missing in logs file, try to append on top
        xlsxFilter.line_prepender(self, filePath, "\n\n#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*\n\n")
        xlsxFilter.line_prepender(self, filePath, str(df))
        xlsxFilter.line_prepender(self, filePath, "Output Length:"+str(len(df))+'\n\n\n')

        if(exportFlag == 1):
            #exportToExcel(xlsxFilePath, df, 'Sheet1')
            demox = df[primarycolumn].value_counts().to_frame()
            #xlsxFilePath = str(filePath).replace('.txt', '.xlsx')
            #xlsxFilePath = str(xlsxFilePath).replace('/Logs', '/Jupyter-Excel-Exports')
            #xlsx = exportToExcel(xlsxFilePath, list([df, demox]))
            xlsx = xlsxFilter.exportToExcel(self, "null", list([df, demox]))
            xlsxFilter.addToLogIndex(self, qCode, queryText, filePath, xlsx, timeStampFileName)
        return df

    def countCrossFields(self, df1, primaryColumn, countColumn, paramsList):
        print("Input Size: ", len(df1))
        print("Get Count wrt Column: ", primaryColumn)
        print("Get Count on Column: ", countColumn)
        print("Search: ", paramsList, "in Column: ", countColumn)
        primaryList = df1[primaryColumn].unique()
        cols = []
        cols.append(primaryColumn)
        cols = cols + paramsList
        df = xlsxFilter.pd.DataFrame(columns=cols)
        df[primaryColumn] = primaryList
        for i in range (0, len(paramsList)):
            df[paramsList[i]] = df[paramsList[i]].fillna(0)
        for i in range (0, len(df1)):
            primarycolData = str(df1[primaryColumn][i])
            columnData = df1[countColumn][i]
            ele, index = xlsxFilter.containList(self, columnData, paramsList)
            if(index != -1):
                colName = paramsList[index]
                temp, dfIndex = xlsxFilter.containList(self, primarycolData, primaryList)
                tempValue = df[colName][dfIndex]
                tempValue += 1
                df[colName][dfIndex] = tempValue
        df = xlsxFilter.removeZeroValueRows(self, df, paramsList)
        print("Output size: ", len(df))
        return df

    def countCrossFieldsNominal(self, df1, primaryColumn, countColumn, paramsList):
        print("Input Size: ", len(df1))
        print("Get Count wrt Column: ", primaryColumn)
        print("Get Count on Column: ", countColumn)
        print("Search: ", paramsList, "in Column: ", countColumn)
        primaryList = df1[primaryColumn].unique()
        cols = []
        cols.append(primaryColumn)
        cols = cols + paramsList
        df = xlsxFilter.pd.DataFrame(columns=cols)
        df[primaryColumn] = primaryList
        for i in range (0, len(paramsList)):
            df[paramsList[i]] = df[paramsList[i]].fillna(0)
        for i in range (0, len(df1)):
            primarycolData = str(df1[primaryColumn][i])
            columnData = df1[countColumn][i]
            if columnData in paramsList:
            #ele, index = containList(columnData, paramsList)
                index = paramsList.index(columnData)
                colName = paramsList[index]
                temp, dfIndex = xlsxFilter.containList(self, primarycolData, primaryList)
                tempValue = df[colName][dfIndex]
                tempValue += 1
                df[colName][dfIndex] = tempValue
                #temp = df1[colName][i]
                #temp += 1
                #df1[columnData][i] = temp
                #if(index != -1):
                #colName = paramsList[index]
        df = xlsxFilter.removeZeroValueRows(self, df, paramsList)
        print("Output size: ", len(df))
        return df

    def transformNominal(self, df1, columnName, newColumnName, paramsList, defaultparam):
        df1[newColumnName] = xlsxFilter.pd.Series(xlsxFilter.np.random.randn(len(df1[columnName])), index=df1.index)
        for index in range (0, len(df1)):
            newValue = ""
            columnValue = df1[columnName][index]
            columnValue = columnValue.lower()
            #print(columnValue)
            newValue, listIndex = xlsxFilter.containList(self, columnValue, paramsList)
            #print(newValue, listIndex)
            if(listIndex == -1):
                newValue = defaultparam
            df1[newColumnName][index] = newValue
        return df1

    def containList(self, searchString, li):
        for i in range (0, len(li)):
            ele = li[i]
            if((str(searchString).find(str(ele)))>-1):
                return [ele, i]
        return [-1, -1]

    def removeZeroValueRows(self, df, cols_of_interest):
        df = df.loc[df[df[cols_of_interest]!=0].dropna(thresh=1).index]
        return df

    def pxExInx(self, df, inclusionList, exclusionList):
        print('Input Size: ', len(df))
        df = df.loc[df[df[exclusionList]==0].dropna(thresh=1).index]
        df = df.loc[df[df[inclusionList]!=0].dropna(thresh=1).index]
        print('Output Size: ', len(df))
        return df

    def findAndConcat(self, df1, primaryColumn, paramsList):
        print('Input Rows:', len(df1))
        colsList = list(df1.columns)
        df = xlsxFilter.pd.DataFrame(columns=colsList)
        for i in range (0, len(paramsList)):
            dfx = df1[df1[primaryColumn]==str(paramsList[i])]
            df = xlsxFilter.pd.concat([df,dfx]).drop_duplicates().reset_index(drop=True)
        print('Output Rows:', len(df))
        return df

    def transformNominalBinary(self, df1, columnName, newColumnName, paramsList):
        df1[newColumnName] = xlsxFilter.pd.Series(xlsxFilter.np.random.randn(len(df1[columnName])), index=df1.index)
        for index in range (0, len(df1)):
            newValue = "False"
            columnValue = df1[columnName][index]
            newValue, listIndex = xlsxFilter.containList(self, columnValue, paramsList)
            if(listIndex == -1):
                newValue = "False"
            else:
                newValue = 'True'
            df1[newColumnName][index] = newValue
        return df1

    def getUniqueCount(self, df1):
        colDict = dict()
        allCols = list(df1.columns)
        for i in range (0, len(allCols)):
            distinctCount = len(df1[allCols[i]].unique())
            #print(allCols[i], ":::", distinctCount)
            colDict[allCols[i]]=distinctCount
        from operator import itemgetter
        print(sorted(colDict.items(), key=itemgetter(1)),)
        return colDict

    def vizualizeDf(self, df):
        import pandas as pd
        import numpy as np
        import cufflinks as cf
       # %matplotlib inline
        from plotly.offline import download_plotlyjs, init_notebook_mode, plot, iplot
        init_notebook_mode(connected=True)
        cf.go_offline()
        import seaborn as sns
        df.iplot(kind='bar')

    def vizualizeDfsegment(df, bin):
        import pandas as pd
        import numpy as np
        import cufflinks as cf
       # %matplotlib inline
        from plotly.offline import download_plotlyjs, init_notebook_mode, plot, iplot
        init_notebook_mode(connected=True)
        cf.go_offline()
        import seaborn as sns
        df.iplot(kind='bar')
        i = 0
        while i <= len(df):
            df[i:i+bin].iplot(kind='bar')
            i = i + bin
